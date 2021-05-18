# 王琰的python编写
# 开发时间:2021/4/22 9:48

from openpyxl import Workbook
from openpyxl import load_workbook
from pandas.io.excel import ExcelWriter
import pandas as pd

# 转格式
# pd.read_csv('D:/我的成长/2021开心的我/生活/股票池/early0422.csv').to_excel('D:/我的成长/2021开心的我/生活/股票池/early0422.xlsx')

# 实例化
wb = Workbook()
# active
ws = wb.active
# 直接读取excel工作簿
wb2 = load_workbook('D:/我的成长/2021开心的我/生活/股票池/517-925early.xlsx')
print(wb2.sheetnames)
# 取工作表
sheet = wb2['Sheet1']
# 表最大行；最大列
max_row = sheet.max_row
print(max_row)
D = []
for m_row in range(max_row-1):
# for m_row in range(3):
    # 读取单元格信息
    d = sheet.cell(row=m_row+2,column=1).value
    # print(d)
    d1 = d.split('\n')
    # print(d1)

    for i in range(len(d1)):
        while i >> 0:
            i1 = str(d1[i])
            print(i1)
            d2 = i1.split(' ')
            print(d2)
            for r in range(len(d2)):
                if d2[r] == '数据':
                    print(r+1)
                    print(d2[0:3])
                    list1 = d2[0:3]
                    print(d2[r+1:])
                    list2 = d2[r+1:]
                    list1.extend(list2)
                    D.append(list1)
                    print(D)
                else:
                    continue

            break

result = pd.DataFrame(D)

result.to_excel('D:/我的成长/2021开心的我/生活/股票池//集合竞价/517-925early8.xlsx')


