# 王琰的python编写
# 开发时间:2021/5/15 8:37

from openpyxl import Workbook
from openpyxl import load_workbook
from pandas.io.excel import ExcelWriter
import pandas as pd
import re
import numpy as np

# 实例化
wb = Workbook()
# active
ws = wb.active
# 直接读取excel工作簿
wb2 = load_workbook('D:/我的成长/2021开心的我/生活/股票池/集合竞价/514-925early8.xlsx')
print(wb2.sheetnames)

# 取工作表
sheet = wb2['Sheet1']
# 表最大行；最大列
max_row = sheet.max_row
print(max_row)

# D = []

for m_row in range(max_row-1):
    # 读取单元格信息
    d = sheet.cell(row=m_row+1,column=9).value
    # print(d)
    pattern = re.search('万',d)
    # pat = pattern.group()
    # print(pattern)
    if pattern is not None:
        pattern1 = re.search('(\d+).*',d)
        pattern2 = pattern1.group()
        pattern3 = pattern2.split('万')
        print(pattern3)
        pattern4 = np.array(pattern3[0])
        print(pattern4)
        dm = float(pattern4) * float(10000)
        print(dm)
        # print((pattern3*10000)[0])
        # D.append((pattern3*10000)[0])
    else:
        dm = d
        # print(dm)
        # D.append(d)
    sheet.cell(row=m_row+1,column=30).value = dm

for m_row in range(max_row-1):
    # 读取单元格信息
    d_1 = sheet.cell(row=m_row+1,column=23).value
    # print(d)
    pattern_1 = re.search('万',d_1)
    # pat = pattern.group()
    # print(pattern)
    if pattern_1 is not None:
        pattern1_1 = re.search('(\d+).*',d_1)
        pattern2_1 = pattern1_1.group()
        pattern3_1 = pattern2_1.split('万')
        print(pattern3_1)
        pattern4_1 = np.array(pattern3_1[0])
        print(pattern4_1)
        dm2 = float(pattern4_1) * float(10000)
        print(dm2)
        # print((pattern3*10000)[0])
        # D.append((pattern3*10000)[0])
    else:
        dm2 = d_1
        # print(dm)
        # D.append(d)
    sheet.cell(row=m_row+1,column=31).value = dm2

wb2.save('D:/我的成长/2021开心的我/生活/股票池/集合竞价/514-925early8.xlsx')